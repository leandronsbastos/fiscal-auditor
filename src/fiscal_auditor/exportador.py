"""
Gerador de relatórios exportáveis (Excel e PDF).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from typing import List, Dict
import json


class ExportadorRelatorios:
    """Classe para exportar relatórios em diferentes formatos."""
    
    @staticmethod
    def gerar_excel(dados_sessao: dict) -> BytesIO:
        """
        Gera arquivo Excel com múltiplas abas contendo todos os relatórios.
        
        Args:
            dados_sessao: Dicionário com documentos, mapa, validações e relatórios
            
        Returns:
            BytesIO com o arquivo Excel
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove planilha padrão
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aba 1: Resumo
        ws_resumo = wb.create_sheet("Resumo")
        ExportadorRelatorios._criar_aba_resumo(ws_resumo, dados_sessao, header_fill, header_font, border)
        
        # Aba 2: Mapa de Apuração
        ws_mapa = wb.create_sheet("Mapa de Apuração")
        ExportadorRelatorios._criar_aba_mapa(ws_mapa, dados_sessao, header_fill, header_font, border)
        
        # Aba 3: Documentos
        ws_docs = wb.create_sheet("Documentos")
        ExportadorRelatorios._criar_aba_documentos(ws_docs, dados_sessao, header_fill, header_font, border)
        
        # Aba 4: Entradas
        ws_entradas = wb.create_sheet("Entradas")
        ExportadorRelatorios._criar_aba_entradas(ws_entradas, dados_sessao, header_fill, header_font, border)
        
        # Aba 5: Saídas
        ws_saidas = wb.create_sheet("Saídas")
        ExportadorRelatorios._criar_aba_saidas(ws_saidas, dados_sessao, header_fill, header_font, border)
        
        # Aba 6: Validações
        ws_validacoes = wb.create_sheet("Validações")
        ExportadorRelatorios._criar_aba_validacoes(ws_validacoes, dados_sessao, header_fill, header_font, border)
        
        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def _criar_aba_resumo(ws, dados, header_fill, header_font, border):
        """Cria aba de resumo."""
        empresa = dados.get("empresa", {})
        mapa = dados.get("mapa")
        documentos = dados.get("documentos", [])
        
        ws['A1'] = "RELATÓRIO DE APURAÇÃO TRIBUTÁRIA"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = "Empresa:"
        ws[f'B{row}'] = empresa.get("razao_social", "N/A")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "CNPJ:"
        ws[f'B{row}'] = empresa.get("cnpj", "N/A")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Período:"
        ws[f'B{row}'] = mapa.periodo if mapa else "N/A"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Data do Relatório:"
        ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 2
        ws[f'A{row}'] = "Total de Documentos:"
        ws[f'B{row}'] = len(documentos)
        ws[f'A{row}'].font = Font(bold=True)
        
        # Contadores
        entradas = sum(1 for d in documentos if d.tipo_movimento.value == "Entrada")
        saidas = sum(1 for d in documentos if d.tipo_movimento.value == "Saída")
        
        row += 1
        ws[f'A{row}'] = "Entradas:"
        ws[f'B{row}'] = entradas
        
        row += 1
        ws[f'A{row}'] = "Saídas:"
        ws[f'B{row}'] = saidas
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    @staticmethod
    def _criar_aba_mapa(ws, dados, header_fill, header_font, border):
        """Cria aba do mapa de apuração."""
        mapa = dados.get("mapa")
        if not mapa:
            ws['A1'] = "Nenhum mapa de apuração disponível"
            return
        
        headers = ["Tributo", "Débitos", "Créditos", "Saldo"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for apuracao in mapa.apuracoes:
            ws.cell(row, 1, apuracao.tipo.value).border = border
            ws.cell(row, 2, float(apuracao.debitos)).border = border
            ws.cell(row, 2).number_format = '#,##0.00'
            ws.cell(row, 3, float(apuracao.creditos)).border = border
            ws.cell(row, 3).number_format = '#,##0.00'
            ws.cell(row, 4, float(apuracao.saldo)).border = border
            ws.cell(row, 4).number_format = '#,##0.00'
            row += 1
        
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    @staticmethod
    def _criar_aba_documentos(ws, dados, header_fill, header_font, border):
        """Cria aba de documentos."""
        documentos = dados.get("documentos", [])
        
        headers = ["Chave", "Número", "Tipo", "Movimento", "Emissão", "CNPJ Emitente", "CNPJ Destinatário", "Valor Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for doc in documentos:
            ws.cell(row, 1, doc.chave).border = border
            ws.cell(row, 2, doc.numero).border = border
            ws.cell(row, 3, doc.tipo.value).border = border
            ws.cell(row, 4, doc.tipo_movimento.value).border = border
            ws.cell(row, 5, doc.data_emissao.strftime("%d/%m/%Y") if doc.data_emissao else "").border = border
            ws.cell(row, 6, doc.cnpj_emitente).border = border
            ws.cell(row, 7, doc.cnpj_destinatario).border = border
            ws.cell(row, 8, float(doc.valor_total)).border = border
            ws.cell(row, 8).number_format = '#,##0.00'
            row += 1
        
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
    
    @staticmethod
    def _criar_aba_entradas(ws, dados, header_fill, header_font, border):
        """Cria aba de documentos de entrada."""
        documentos = [d for d in dados.get("documentos", []) if d.tipo_movimento.value == "Entrada"]
        ExportadorRelatorios._preencher_aba_movimento(ws, documentos, "ENTRADAS", header_fill, header_font, border)
    
    @staticmethod
    def _criar_aba_saidas(ws, dados, header_fill, header_font, border):
        """Cria aba de documentos de saída."""
        documentos = [d for d in dados.get("documentos", []) if d.tipo_movimento.value == "Saída"]
        ExportadorRelatorios._preencher_aba_movimento(ws, documentos, "SAÍDAS", header_fill, header_font, border)
    
    @staticmethod
    def _preencher_aba_movimento(ws, documentos, titulo, header_fill, header_font, border):
        """Preenche aba de movimento (entrada ou saída)."""
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        
        headers = ["Documento", "Data", "Fornecedor/Cliente", "Valor", "ICMS", "IPI", "PIS", "COFINS", "IBS", "CBS"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row = 4
        for doc in documentos:
            ws.cell(row, 1, f"{doc.numero}/{doc.serie}").border = border
            ws.cell(row, 2, doc.data_emissao.strftime("%d/%m/%Y") if doc.data_emissao else "").border = border
            
            # Fornecedor ou cliente dependendo do movimento
            cnpj = doc.cnpj_emitente if titulo == "ENTRADAS" else doc.cnpj_destinatario
            ws.cell(row, 3, cnpj).border = border
            
            ws.cell(row, 4, float(doc.valor_total)).border = border
            ws.cell(row, 4).number_format = '#,##0.00'
            
            # Somar tributos do documento
            tributos_totais = {"ICMS": 0, "IPI": 0, "PIS": 0, "COFINS": 0, "IBS": 0, "CBS": 0}
            for item in doc.items:
                for tributo in item.tributos:
                    if tributo.tipo.value in tributos_totais:
                        tributos_totais[tributo.tipo.value] += float(tributo.valor)
            
            col = 5
            for tipo_trib in ["ICMS", "IPI", "PIS", "COFINS", "IBS", "CBS"]:
                ws.cell(row, col, tributos_totais[tipo_trib]).border = border
                ws.cell(row, col).number_format = '#,##0.00'
                col += 1
            
            row += 1
        
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    @staticmethod
    def _criar_aba_validacoes(ws, dados, header_fill, header_font, border):
        """Cria aba de validações."""
        validacoes = dados.get("validacoes", [])
        
        headers = ["Documento", "Status", "Mensagens", "Créditos Aproveitáveis", "Créditos Indevidos", "Créditos Glosáveis"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row = 2
        for val in validacoes:
            ws.cell(row, 1, val.chave_acesso).border = border
            ws.cell(row, 2, "Válido" if val.valido else "Com Problemas").border = border
            ws.cell(row, 3, len(val.mensagens)).border = border
            ws.cell(row, 4, len(val.creditos_aproveitaveis)).border = border
            ws.cell(row, 5, len(val.creditos_indevidos)).border = border
            ws.cell(row, 6, len(val.creditos_glosaveis)).border = border
            row += 1
            
            # Detalhar mensagens
            if val.mensagens:
                for msg in val.mensagens:
                    ws.cell(row, 2, msg).border = border
                    ws['B' + str(row)].font = Font(color="FF6600", italic=True)
                    row += 1
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
    
    @staticmethod
    def gerar_pdf(dados_sessao: dict) -> BytesIO:
        """
        Gera relatório em PDF.
        
        Args:
            dados_sessao: Dicionário com documentos, mapa, validações
            
        Returns:
            BytesIO com o arquivo PDF
        """
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=2*cm, rightMargin=2*cm)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Estilo customizado
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Título
        empresa = dados_sessao.get("empresa", {})
        mapa = dados_sessao.get("mapa")
        
        story.append(Paragraph("RELATÓRIO DE APURAÇÃO TRIBUTÁRIA", title_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Cabeçalho
        info_data = [
            ["Empresa:", empresa.get("razao_social", "N/A")],
            ["CNPJ:", empresa.get("cnpj", "N/A")],
            ["Período:", mapa.periodo if mapa else "N/A"],
            ["Data:", datetime.now().strftime("%d/%m/%Y %H:%M")]
        ]
        
        info_table = Table(info_data, colWidths=[5*cm, 15*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 1*cm))
        
        # Mapa de Apuração
        if mapa:
            story.append(Paragraph("Mapa de Apuração", styles['Heading2']))
            story.append(Spacer(1, 0.3*cm))
            
            mapa_data = [['Tributo', 'Débitos', 'Créditos', 'Saldo']]
            for apuracao in mapa.apuracoes:
                mapa_data.append([
                    apuracao.tipo.value,
                    f"R$ {float(apuracao.debitos):,.2f}",
                    f"R$ {float(apuracao.creditos):,.2f}",
                    f"R$ {float(apuracao.saldo):,.2f}"
                ])
            
            mapa_table = Table(mapa_data, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
            mapa_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            story.append(mapa_table)
            story.append(PageBreak())
        
        # Resumo de Documentos
        documentos = dados_sessao.get("documentos", [])
        entradas = [d for d in documentos if d.tipo_movimento.value == "Entrada"]
        saidas = [d for d in documentos if d.tipo_movimento.value == "Saída"]
        
        story.append(Paragraph("Resumo de Documentos", styles['Heading2']))
        story.append(Spacer(1, 0.3*cm))
        
        resumo_data = [
            ["Total de Documentos", str(len(documentos))],
            ["Documentos de Entrada", str(len(entradas))],
            ["Documentos de Saída", str(len(saidas))],
        ]
        
        resumo_table = Table(resumo_data, colWidths=[8*cm, 5*cm])
        resumo_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ]))
        story.append(resumo_table)
        
        # Construir PDF
        doc.build(story)
        output.seek(0)
        return output
